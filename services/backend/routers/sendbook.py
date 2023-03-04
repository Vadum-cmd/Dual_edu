import shutil
from typing import List

from openpyxl import load_workbook
from fastapi import FastAPI, Depends, UploadFile, File, Form
from sqlalchemy.orm import Session
from dependencies import get_db
from logic.extract_words_and_translate import extract_text_from_pdf, word_tokenization, translate_word
from main import app
from crud.crud_functions import UserWordCreate, create_user_word, get_db_word


# TODO: rewrite this function
from schemas.new_schema import DBWord


@app.post("/sendbook")
def post_book(file: UploadFile = File(...), level: str = "a2", db: Session = Depends(get_db)):# -> List[DBWord]:
    with open(f'{file.filename}', "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    text = extract_text_from_pdf(file.filename)
    words = word_tokenization(text)
    user_words = []
    book_id = 4
    # for word in words:
    #     create_user_word(db=db, user_word=UserWordCreate(en_word=word, is_known=False, book_id=book_id))
    #
    # # TODO: what about translation
    # for word in words:
    #     user_words.append(get_db_word(db=db, en_word=word))
    #
    # return user_words
    return {"words": "words"}

    # split text from book
    # write words to DB using level
    # return words to user

    # user_words = {}
    # j = 0
    # for word in words:
    #     user_words[word] = f"Error"
    #     j += 1
    #     if j == 50:
    #         break
    #
    # wb = load_workbook(filename='words_table.xlsx', data_only=True)
    # # data_only=False by default
    # # If you want to see data instead of formulas, set data_only=True
    #
    # ws = wb['Аркуш1']  # Or whatever your sheet name is
    # #
    # # translated_text = translator.translate('안녕하세요.')
    # # print(translated_text.text)
    # #
    # # print(translated_text.text)
    # print("OK by now")
    # for us_word in user_words.keys():
    #     for num_row in range(1, ws.max_row):
    #         if ws[f'A{num_row}'].value == us_word:
    #             try:
    #                 user_words[us_word] = translate_word(us_word)
    #             except:
    #                 print(us_word)
    # print(user_words)
    #return {"words": user_words}
